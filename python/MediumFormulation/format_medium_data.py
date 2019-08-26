"""
Extracting data from CCLE data obtained from Ghandi et al., 2019
The files that are needed as the input for the script:
  * CCLE_GCP.csv - obtained from the supplementary data under a different name, but it is the file that contains the global chromatin proteomics dataset
The files created from this script:
  * ccle_names.txt - the names of the CCLE cell lines corresponding to the H3 markers
  * h3marks.txt - the names of the H3 markers
  * h3_relval.txt - the values of the relative proteomics intensities obtained from Ghandi et al., 2019.
@author: Scott Campit
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import chain
import string


def get_media(dictionary, key, lst):
    """
    This function requires a list of synonyms for a specific medium. It will construct a dictionary that maps the key to various synonyms of the medium.
    """
    for i in lst:
        try:
            dictionary[key].append(i)
        except KeyError:
            dictionary[key] = [i]
    return dictionary


def mapper(dict, series):
    """
    This function maps the values in a medium key to the dataframe elements to get the keys into the dataframe.
    """
    for k, v in dict.items():
        idx = series.isin(v)
        tmp = series[idx]
        tmp[idx] = k
        series[idx] = tmp
    return series


def split_cellLine_and_tissues():
    df = pd.read_csv('GCP_proteomics_remapped.csv')
    df['Cell Line'] = df['CellLineName'].str.split('_').str[0]
    df['Tissue'] = df['CellLineName'].str.split('_', n=1).str[1]
    df['Tissue'] = df['Tissue'].str.replace('_', ' ')
    df['Tissue'] = df['Tissue'].str.title()
    df = df.drop(['CellLineName'], axis=1)
    df = df.set_index(['Cell Line', 'Tissue'])

    df.to_csv("GCP_proteomics_remapped.csv")
#split_cellLine_and_tissues()


def make_medium_xl_sheet():
    """
    """

    medium_conditions = pd.read_csv('GCP_proteomics_remapped.csv', usecols=[
                                    'Medium Condition', 'Tissue'])
    unique_conditions = medium_conditions['Medium Condition'].sort_values(
        ascending=True).unique()

    #wb = Workbook()
    #name = 'Medium_conditions.xlsx'
    #wb.save(filename = name)

    number_of_medium = medium_conditions['Medium Condition'].value_counts()
    number_of_tissues = medium_conditions['Tissue'].value_counts()
    book = load_workbook(
        r'./../../data/Medium_Component_Maps/final_medium.xlsx')
    writer = pd.ExcelWriter(
        r'./../../data/Medium_Component_Maps/final_medium.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    if "Summary" in book:
        pass
    else:
        number_of_medium.to_excel(
            writer, sheet_name="Summary", index=True, startcol=0)
        number_of_tissues.to_excel(
            writer, sheet_name="Summary", index=True, startcol=2)
    writer.save()

    for medium in unique_conditions:
        if medium in book:
            pass
        else:
            df = pd.DataFrame(
                columns=["Components", "MW", "g/L", "mM", "BiGG ID", "Alpha", "LB", "Adjusted LB"])
        df.to_excel(writer, sheet_name=medium, index=False, header=True)
    writer.save()
#make_medium_xl_sheet()


def get_unique_ids():
    df = pd.read_csv('./medium_component_map.csv')
    df = df.drop_duplicates(keep='first')
    df.to_csv('./medium_component_map.csv', index=False)
#get_unique_ids()


def map_to_recon1_xchange():
    pd.options.mode.chained_assignment = None
    df = pd.read_csv('./medium_component_map.csv')
    metabolites = pd.read_excel('./recon1_id.xlsx', sheet_name='Metabolites')
    reactions = pd.read_excel('./recon1_id.xlsx', sheet_name='Reactions')

    exchange_reactions = reactions[reactions['Reaction ID'].str.startswith(
        'EX_')]
    exchange_reactions['Metabolite'] = exchange_reactions['Reaction ID'].str.replace(
        'EX_', '')

    recon1_map = pd.merge(exchange_reactions, metabolites,
                          how='inner', left_on='Metabolite', right_on='Metabolite ID')
    recon1_map = recon1_map.copy(deep=True)
    full_map = pd.merge(recon1_map, df, how='inner',
                        left_on='Metabolite Name', right_on='BiGG Common Names')
    full_map = full_map.drop_duplicates(keep='first')
    full_map.to_csv('./medium_component_mapped_to_recon1.csv', index=False)
    #recon1_map.to_csv('Recon1_exchange_map.csv', index=False)


#map_to_recon1_xchange()

def recon_xchange():
    pd.options.mode.chained_assignment = None

    metabolites = pd.read_excel(
        './../../data/metabolicModel_maps/RECON3_Reaction_Metabolite_ID_Maps.xlsx', sheet_name='Metabolites')
    reactions = pd.read_excel(
        './../../data/metabolicModel_maps/RECON3_Reaction_Metabolite_ID_Maps.xlsx', sheet_name='Reactions')

    exchange_reactions = reactions[reactions['Bigg Reaction ID'].str.startswith(
        'EX_')]
    exchange_reactions['Metabolite'] = exchange_reactions['Bigg Reaction ID'].str.replace(
        'EX_', '')

    recon1_map = pd.merge(exchange_reactions, metabolites,
                          how='inner', left_on='Metabolite', right_on='BiGG Metabolite ID')
    recon1_map = recon1_map.copy(deep=True)
    recon1_map.to_csv(
        './../../data/metabolicModel_maps/RECON3_ExchangeReaction_Map.csv', index=False)


recon_xchange()


def drop_all_duplicates():
    df = pd.read_excel(r'./../../data/Medium_Component_Maps/final_medium.xlsx')
    df = df.drop_duplicates(keep='first')
    df.to_excel(
        r'./../../data/Medium_Component_Maps/final_medium.xlsx', index=False)


def average_all_duplicates(df):
    df = df.groupby(df.index).mean().reset_index()
    df = df.set_index("Components")
    return df


#drop_all_duplicates()


def make_medium_conditions():

    rpmi = pd.read_excel(
        r'./../../data/Medium_Component_Maps/original_medium.xlsx', sheet_name='RPMI')
    dmem = pd.read_excel(
        r'./../../data/Medium_Component_Maps/original_medium.xlsx', sheet_name='DMEM')
    imdm = pd.read_excel(
        r'./../../data/Medium_Component_Maps/original_medium.xlsx', sheet_name='IMDM')
    emem = pd.read_excel(
        r'./../../data/Medium_Component_Maps/original_medium.xlsx', sheet_name='EMEM')
    mcdb105 = pd.read_excel(
        r'./../../data/Medium_Component_Maps/original_medium.xlsx', sheet_name='MCDB105')
    m199 = pd.read_excel(
        r'./../../data/Medium_Component_Maps/original_medium.xlsx', sheet_name='M199')
    f12 = pd.read_excel(
        r'./../../data/Medium_Component_Maps/original_medium.xlsx', sheet_name='F12')

    book = load_workbook(
        r'./../../data/Medium_Component_Maps/final_medium.xlsx')
    writer = pd.ExcelWriter(
        r'./../../data/Medium_Component_Maps/final_medium.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    def make_medium(medium1, medium2, weight=0.5):
        """
        args:
            weight: the proportionality constant (default = 1:1)
        """

        weight1 = weight
        weight2 = 1 - weight

        medium1 = medium1.copy(deep=True)
        medium1['g/L'] = medium1['g/L'].replace('Infinity', np.inf)
        medium1['g/L'] = medium1['g/L']*weight1

        medium2 = medium2.copy(deep=True)
        medium2['g/L'] = medium2['g/L'].replace('Infinity', np.inf)
        medium2['g/L'] = medium2['g/L']*weight2

        combined_medium = pd.concat([medium1, medium2]).groupby(
            'Components')['g/L'].sum().reset_index()

        rest_of_columns = ["MW", "mM", "BiGG ID",
                           "Alpha", "LB", "Adjusted LB"]
        combined_medium = pd.concat(
            [combined_medium, pd.DataFrame(columns=rest_of_columns)], sort=False)

        return combined_medium

    # DMEM-IMDM
    dmem_imdm = make_medium(dmem, imdm)
    dmem_imdm.to_excel(writer, sheet_name='DMEM-IMDM', columns=[
                       "Components", "MW", "g/L", "mM", "BiGG ID", "Alpha", "LB", "Adjusted LB"], index=False)

    # MCDB105-M199
    mcdb105_m199 = make_medium(mcdb105, m199)
    mcdb105_m199.to_excel(writer, sheet_name='MCDB105-M199', index=False)

    # RPMI-EMEM
    rpmi_emem = make_medium(rpmi, emem)
    rpmi_emem.to_excel(writer, sheet_name='RPMI-EMEM', index=False)

    # RPMI-F12
    rpmi_f12 = make_medium(rpmi, f12)
    rpmi_f12.to_excel(writer, sheet_name='RPMI-F12', index=False)

    dmem_rpmi = make_medium(dmem, rpmi, weight=2/3)
    rpmi_f12.to_excel(writer, sheet_name='DMEM-RPMI', index=False)

    writer.save()


#make_medium_conditions()


def calculate_alpha():

    book = load_workbook(
        r'./../../data/Medium_Component_Maps/final_medium.xlsx')
    writer = pd.ExcelWriter(
        r'./../../data/Medium_Component_Maps/final_medium.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    rpmi = pd.read_excel(r'./../../data/Medium_Component_Maps/final_medium.xlsx',
                         sheet_name='RPMI', index_col='Components')
    rpmi = average_all_duplicates(rpmi)
    rpmi["LB"] = -1

    for medium in writer.sheets:
        if medium == "Summary":
            pass
        else:
            other_medium = pd.read_excel(
                r'./../../data/Medium_Component_Maps/final_medium.xlsx', sheet_name=medium, index_col='Components')
            other_medium = average_all_duplicates(other_medium)
            other_medium["LB"] = -1
            other_medium['alpha'] = other_medium['g/L'].divide(
                rpmi['g/L'], axis=0, fill_value=0)
            other_medium['alpha'] = other_medium['alpha'].replace(np.inf, 10)
            other_medium.to_excel(writer, sheet_name=medium, index=True)
    writer.save()


#calculate_alpha()


def map_recon1_xchange_to_medium():
    book = load_workbook(
        r'./../../data/Medium_Component_Maps/final_medium.xlsx')
    writer = pd.ExcelWriter(
        r'./../../data/Medium_Component_Maps/final_medium2.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    metabolite_map = pd.read_csv(
        './../../data/metabolicModel_maps/RECON3_ExchangeReaction_Map.csv')

    for medium in writer.sheets:
        if medium == "Summary":
            pass
        else:
            df = pd.read_excel(
                r'./../../data/Medium_Component_Maps/final_medium.xlsx', sheet_name=medium)
            merged_df = pd.merge(
                df, metabolite_map, how='inner', left_on=['Components'], right_on=['Metabolite Name'])
            merged_df = merged_df.drop_duplicates(keep='first')
            merged_df.to_excel(
                writer, sheet_name=medium, index=False, header=True)
    writer.save()


#map_recon1_xchange_to_medium()


def scale_LB():
    # load the excel file so you don't overwrite the excel sheet
    book = load_workbook(
        r'./../../data/Medium_Component_Maps/final_medium2.xlsx')
    writer = pd.ExcelWriter(
        r'./../../data/Medium_Component_Maps/final_medium3.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    default_uptake_rate = pd.read_excel(
        r'./../../data/Medium_Component_Maps/final_medium2.xlsx', sheet_name='RPMI', index_col='Components')

    for medium in writer.sheets:
        if medium == "Summary":
            pass
        else:
            uptake_rate_to_change = pd.read_excel(
                r'./../../data/Medium_Component_Maps/final_medium2.xlsx', sheet_name=medium, index_col='Components')

            uptake_rate_to_change['Adjusted LB'] = default_uptake_rate['LB'].multiply(
                uptake_rate_to_change['alpha'], axis=0, fill_value=1.00)
            uptake_rate_to_change.to_excel(
                writer, sheet_name=medium, index=True)
    writer.save()


#scale_LB()